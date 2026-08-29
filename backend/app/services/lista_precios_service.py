import re
import uuid
from decimal import Decimal
from io import BytesIO

from fastapi.exceptions import RequestValidationError
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.core.database import utcnow
from app.exceptions.base import BadRequestError, ConflictError, NotFoundError
from app.models.articulo import Articulo
from app.models.lista_precios import ListaPrecios
from app.models.medida import Medida
from app.models.proveedor import Proveedor
from app.models.proveedor_categoria import ProveedorCategoria
from app.repositories.articulo_repository import ArticuloRepository
from app.repositories.categoria_repository import CategoriaRepository
from app.repositories.lista_precios_repository import ListaPreciosRepository
from app.repositories.medida_repository import MedidaRepository
from app.repositories.proveedor_categoria_repository import ProveedorCategoriaRepository
from app.repositories.proveedor_repository import ProveedorRepository
from app.schemas.alta_inventario import ArticuloAlta, MedidaAlta
from app.schemas.lista_precios import (
    ItemListaPrecio,
    ListaPreciosAlta,
    ListaPreciosExcelAlta,
    ListaPreciosUpdate,
    MapeoColumna,
    ProveedorAlta,
)

VALID_MAPEO_KEYS = {
    "articulo_id",
    "nombre",
    "articulo_medida_combinado",
    "descripcion",
    "categoria",
    "medida_id",
    "unidad_medida",
    "medida",
    "id_articulo_proveedor",
    "precio_lista",
}

KEYS_INCOMPATIBLES_COMBINADO = {"articulo_id", "nombre", "unidad_medida", "medida", "medida_id"}

UNIDADES_CONOCIDAS = {
    "kg": "kg",
    "g": "g",
    "l": "l",
    "ml": "ml",
    "lt": "l",
    "lts": "l",
    "m": "m",
    "mt": "m",
    "mts": "m",
    "cm": "cm",
    "mm": "mm",
    "un": "unidad",
    "u": "unidad",
    "unidad": "unidad",
    "unidades": "unidad",
    "pz": "pieza",
    "pieza": "pieza",
    "piezas": "pieza",
}

MEDIDA_NO_CORRESPONDE = ("no corresponde", "no corresponde")

_PATRON_CANTIDAD_UNIDAD = re.compile(r"(?P<cantidad>\d+(?:[.,]\d+)?)\s*(?P<unidad>[a-zA-ZñÑ]+)")

_PATRON_MONEDA = re.compile(r"^(?:usd|u\$s|ar\$|\$)\s*", re.IGNORECASE)

_PATRON_MILES = re.compile(r"^\d{1,3}(?:\.\d{3})+$")


def _normalizar_precio(valor: str) -> str:
    texto = _PATRON_MONEDA.sub("", valor.strip()).replace(" ", "")
    if "," in texto and "." in texto:
        if texto.rfind(",") > texto.rfind("."):
            return texto.replace(".", "").replace(",", ".")
        return texto.replace(",", "")
    if "," in texto:
        return texto.replace(",", ".")
    if _PATRON_MILES.match(texto):
        return texto.replace(".", "")
    return texto


def _parsear_articulo_medida(texto: str) -> tuple[str, str, str]:
    normalizado = " ".join(texto.split())
    match = None
    for candidato in _PATRON_CANTIDAD_UNIDAD.finditer(normalizado):
        if candidato.group("unidad").lower() in UNIDADES_CONOCIDAS:
            match = candidato
    if match is None:
        return normalizado, *MEDIDA_NO_CORRESPONDE
    unidad = UNIDADES_CONOCIDAS[match.group("unidad").lower()]
    cantidad = match.group("cantidad").replace(",", ".")
    nombre = " ".join((normalizado[: match.start()] + " " + normalizado[match.end() :]).split())
    return nombre or normalizado, unidad, cantidad


def _validation_error(field: str, msg: str) -> RequestValidationError:
    return RequestValidationError(errors=[{"loc": ("body", field), "msg": msg, "type": "value_error"}])


def _error_celda(fila: int, columna: str, msg: str) -> RequestValidationError:
    return RequestValidationError(
        errors=[{"loc": ("body", f"fila {fila}", f"columna '{columna}'"), "msg": msg, "type": "value_error"}]
    )


class ListaPreciosService:
    def __init__(self, db: Session):
        self.db = db
        self.repository = ListaPreciosRepository(db)
        self.proveedor_repository = ProveedorRepository(db)
        self.proveedor_categoria_repository = ProveedorCategoriaRepository(db)
        self.articulo_repository = ArticuloRepository(db)
        self.medida_repository = MedidaRepository(db)
        self.categoria_repository = CategoriaRepository(db)

    def list(
        self,
        skip: int = 0,
        limit: int = 100,
        categoria_ids: list[uuid.UUID] | None = None,
        articulo_ids: list[uuid.UUID] | None = None,
        proveedor_id: uuid.UUID | None = None,
    ) -> list[ListaPrecios]:
        return self.repository.list(
            skip=skip,
            limit=limit,
            categoria_ids=categoria_ids,
            articulo_ids=articulo_ids,
            proveedor_id=proveedor_id,
        )

    def get(self, lista_precios_id: uuid.UUID) -> ListaPrecios:
        lista = self.repository.get(lista_precios_id)
        if lista is None:
            raise NotFoundError(detail="Registro de lista de precios no encontrado")
        return lista

    def cantidad_por_proveedor(self) -> list[dict]:
        totales = self.repository.contar_por_proveedor()
        desglose_rows = self.repository.contar_por_proveedor_y_categoria()
        cats_por_prov: dict[uuid.UUID, list[dict]] = {}
        for prov, cat, c in desglose_rows:
            cats_por_prov.setdefault(prov.id, []).append({"categoria": cat, "cantidad": c})
        return [
            {"proveedor": prov, "cantidad": total, "por_categoria": cats_por_prov.get(prov.id, [])}
            for prov, total in totales
        ]

    def alta_json(self, data: ListaPreciosAlta) -> list[ListaPrecios]:
        try:
            proveedor = self._resolve_proveedor(data.proveedor_id, data.proveedor)
            result = self._procesar_items(proveedor, data.items)
            self.repository.commit()
            return result
        except Exception:
            self.db.rollback()
            raise

    def alta_excel(self, data: ListaPreciosExcelAlta, archivo: bytes) -> tuple[list[ListaPrecios], list[dict]]:
        try:
            proveedor = self._resolve_proveedor(data.proveedor_id, data.proveedor)
            filas = self._leer_excel(archivo, data.mapeo)
            items: list[ItemListaPrecio] = []
            descartadas: list[dict] = []
            for numero_fila, fila in filas:
                item, motivo = self._item_desde_fila(numero_fila, fila, data.mapeo)
                if item is None:
                    descartadas.append({"fila": numero_fila, "motivo": motivo})
                else:
                    items.append(item)
            result = self._procesar_items(proveedor, items)
            self.repository.commit()
            return result, descartadas
        except Exception:
            self.db.rollback()
            raise

    def update(self, lista_precios_id: uuid.UUID, data: ListaPreciosUpdate) -> ListaPrecios:
        lista = self.get(lista_precios_id)
        if data.id_articulo_proveedor is not None and lista.id_articulo_proveedor != data.id_articulo_proveedor:
            raise NotFoundError(detail="No existe un registro con ese id_articulo_proveedor")
        lista.precio_lista = data.precio_lista
        return self.repository.update(lista)

    def delete(self, lista_precios_id: uuid.UUID) -> None:
        lista = self.get(lista_precios_id)
        lista.deleted_at = utcnow()
        self.repository.soft_delete(lista)

    def _resolve_proveedor(self, proveedor_id: uuid.UUID | None, proveedor: ProveedorAlta | None) -> Proveedor:
        if proveedor_id is not None:
            existente = self.proveedor_repository.get(proveedor_id)
            if existente is None:
                raise BadRequestError(detail="El proveedor no existe o está eliminado")
            return existente
        if proveedor is None or proveedor.id is not None:
            raise BadRequestError(detail="Debe enviarse proveedor_id o un objeto proveedor sin id")
        if self.proveedor_repository.get_by_telefono(proveedor.telefono) is not None:
            raise ConflictError(detail="Ya existe un proveedor con ese teléfono")
        if self.proveedor_repository.get_by_nombre_apellido(proveedor.nombre, proveedor.apellido) is not None:
            raise ConflictError(detail="Ya existe un proveedor con ese nombre y apellido")
        for categoria_id in proveedor.categoria_ids:
            if self.categoria_repository.get(categoria_id) is None:
                raise BadRequestError(detail="La categoría no existe o está eliminada")
        nuevo = Proveedor(
            nombre=proveedor.nombre,
            apellido=proveedor.apellido,
            telefono=proveedor.telefono,
            direccion=proveedor.direccion,
        )
        self.proveedor_repository.add_flush(nuevo)
        for categoria_id in proveedor.categoria_ids:
            self.proveedor_categoria_repository.add_flush(
                ProveedorCategoria(proveedor_id=nuevo.id, categoria_id=categoria_id)
            )
        return nuevo

    def _resolve_articulo(self, data: ArticuloAlta) -> Articulo:
        if data.id is not None:
            articulo = self.articulo_repository.get(data.id)
            if articulo is None:
                raise BadRequestError(detail="El artículo no existe o está eliminado")
            return articulo
        existente = self.articulo_repository.get_by_nombre(data.nombre)
        if existente is not None:
            return existente
        if data.categoria_id is not None and self.categoria_repository.get(data.categoria_id) is None:
            raise BadRequestError(detail="La categoría no existe o está eliminada")
        articulo = Articulo(
            nombre=data.nombre,
            descripcion=data.descripcion,
            categoria_id=data.categoria_id,
        )
        return self.articulo_repository.add_flush(articulo)

    def _resolve_medida(self, data: MedidaAlta) -> Medida:
        if data.id is not None:
            medida = self.medida_repository.get(data.id)
            if medida is None:
                raise BadRequestError(detail="La medida no existe o está eliminada")
            return medida
        existente = self.medida_repository.get_by_combinacion(data.unidad_medida, data.medida)
        if existente is not None:
            return existente
        medida = Medida(unidad_medida=data.unidad_medida, medida=data.medida)
        return self.medida_repository.add_flush(medida)

    def _procesar_items(self, proveedor: Proveedor, items: list[ItemListaPrecio]) -> list[ListaPrecios]:
        result: list[ListaPrecios] = []
        for item in items:
            articulo = self._resolve_articulo(item.articulo)
            medida = self._resolve_medida(item.medida)
            existente = self.repository.get_by_combinacion(proveedor.id, articulo.id)
            if existente is not None:
                existente.deleted_at = None
                existente.medida_id = medida.id
                existente.id_articulo_proveedor = item.id_articulo_proveedor
                existente.precio_lista = item.precio_lista
                result.append(existente)
                continue
            registro = ListaPrecios(
                articulo_id=articulo.id,
                medida_id=medida.id,
                proveedor_id=proveedor.id,
                id_articulo_proveedor=item.id_articulo_proveedor,
                precio_lista=item.precio_lista,
            )
            result.append(self.repository.add_flush(registro))
        return result

    def _leer_excel(self, archivo: bytes, mapeo: list[MapeoColumna]) -> list[tuple[int, dict]]:
        invalidas = [m.key for m in mapeo if m.key not in VALID_MAPEO_KEYS]
        if invalidas:
            raise _validation_error("mapeo", f"Claves de mapeo no soportadas: {', '.join(invalidas)}")
        keys = {m.key for m in mapeo}
        if "articulo_medida_combinado" in keys:
            conflictivas = sorted(keys & KEYS_INCOMPATIBLES_COMBINADO)
            if conflictivas:
                raise _validation_error(
                    "mapeo",
                    "La clave 'articulo_medida_combinado' no puede combinarse con: " + ", ".join(conflictivas),
                )
        try:
            workbook = load_workbook(BytesIO(archivo), data_only=True)
        except Exception:
            raise BadRequestError(detail="El archivo Excel es inválido") from None
        try:
            hoja = workbook.active
            filas = list(hoja.iter_rows(values_only=True))
        finally:
            workbook.close()
        if not filas:
            raise BadRequestError(detail="El archivo Excel está vacío")
        col_by_name = {str(valor).strip(): indice for indice, valor in enumerate(filas[0]) if valor is not None}
        result: list[tuple[int, dict]] = []
        for numero_fila, raw in enumerate(filas[1:], start=2):
            fila: dict = {}
            for m in mapeo:
                col_idx = col_by_name.get(m.value.strip())
                if col_idx is not None and col_idx < len(raw) and raw[col_idx] is not None:
                    fila[m.key] = raw[col_idx]
            if not fila:
                continue
            result.append((numero_fila, fila))
        if not result:
            raise BadRequestError(detail="El archivo Excel no contiene líneas de datos")
        return result

    def _item_desde_fila(
        self, numero_fila: int, fila: dict, mapeo: list[MapeoColumna]
    ) -> tuple[ItemListaPrecio | None, str | None]:
        col = {m.key: m.value for m in mapeo}

        def error(msg: str, key: str):
            raise _error_celda(numero_fila, col.get(key) or key, msg) from None

        def articulo_nuevo(nombre: str) -> ArticuloAlta:
            categoria_id = None
            valor_categoria = fila.get("categoria")
            if valor_categoria is not None and str(valor_categoria).strip():
                categoria = self.categoria_repository.get_by_nombre(str(valor_categoria).strip())
                if categoria is None:
                    error("la categoría no existe", "categoria")
                categoria_id = categoria.id
            descripcion = None
            if fila.get("descripcion") is not None:
                descripcion = str(fila["descripcion"]).strip()
            return ArticuloAlta(nombre=nombre, descripcion=descripcion, categoria_id=categoria_id)

        combinado = None
        texto_combinado = fila.get("articulo_medida_combinado")
        if texto_combinado is not None and str(texto_combinado).strip():
            combinado = _parsear_articulo_medida(str(texto_combinado))

        articulo = None
        if "articulo_id" in fila:
            try:
                articulo = ArticuloAlta(id=uuid.UUID(str(fila["articulo_id"])))
            except ValueError:
                error("id de artículo inválido", "articulo_id")
        elif combinado is not None:
            nombre, _, _ = combinado
            existente = self.articulo_repository.get_by_nombre(nombre)
            if existente is not None:
                articulo = ArticuloAlta(id=existente.id)
            else:
                articulo = articulo_nuevo(nombre)
        elif "nombre" in fila and str(fila["nombre"]).strip():
            nombre = str(fila["nombre"]).strip()
            existente = self.articulo_repository.get_by_nombre(nombre)
            if existente is not None:
                articulo = ArticuloAlta(id=existente.id)
            else:
                articulo = articulo_nuevo(nombre)

        if articulo is None:
            return None, "falta la identificación del artículo"

        medida = None
        if "medida_id" in fila:
            try:
                medida = MedidaAlta(id=uuid.UUID(str(fila["medida_id"])))
            except ValueError:
                error("id de medida inválido", "medida_id")
        elif combinado is not None:
            _, unidad, cantidad = combinado
            existente = self.medida_repository.get_by_combinacion(unidad, cantidad)
            if existente is not None:
                medida = MedidaAlta(id=existente.id)
            else:
                medida = MedidaAlta(unidad_medida=unidad, medida=cantidad)
        elif "unidad_medida" in fila and "medida" in fila:
            unidad = str(fila["unidad_medida"]).strip()
            texto_medida = str(fila["medida"]).strip()
            if unidad and texto_medida:
                existente = self.medida_repository.get_by_combinacion(unidad, texto_medida)
                if existente is not None:
                    medida = MedidaAlta(id=existente.id)
                else:
                    medida = MedidaAlta(unidad_medida=unidad, medida=texto_medida)

        if medida is None:
            return None, "falta la medida"

        valor_precio = fila.get("precio_lista")
        if valor_precio is None:
            return None, "falta el precio de lista"
        try:
            precio = Decimal(_normalizar_precio(str(valor_precio)))
        except Exception:
            error("precio de lista inválido", "precio_lista")
        if precio < 0:
            error("el precio de lista debe ser mayor o igual a 0", "precio_lista")

        id_articulo_proveedor = None
        if fila.get("id_articulo_proveedor") is not None:
            id_articulo_proveedor = str(fila["id_articulo_proveedor"]).strip()

        item = ItemListaPrecio(
            articulo=articulo,
            medida=medida,
            id_articulo_proveedor=id_articulo_proveedor,
            precio_lista=precio,
        )
        return item, None
